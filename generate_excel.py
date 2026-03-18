"""
Generate Excel sheet: Backend → Calculation Engine Data Reference
"""
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ─────────────────────────────────────────────
# COLORS
# ─────────────────────────────────────────────
CLR_HEADER_DARK  = "1F3864"   # dark navy
CLR_HEADER_MID   = "2E75B6"   # blue
CLR_SECTION_A    = "D6E4F7"   # light blue  (backend data)
CLR_SECTION_B    = "E2EFDA"   # light green (user input)
CLR_SECTION_C    = "FFF2CC"   # yellow      (calculated by engine)
CLR_SECTION_D    = "FCE4D6"   # orange      (PHREEQC output)
CLR_ROW_ALT      = "F5F5F5"
CLR_WHITE        = "FFFFFF"
CLR_GREEN_LABEL  = "375623"
CLR_BLUE_LABEL   = "1F3864"
CLR_ORANGE_LABEL = "833C00"
CLR_PURPLE_LABEL = "44136E"

thin  = Side(style="thin",   color="BFBFBF")
thick = Side(style="medium", color="1F3864")

def hdr_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def row_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def hdr_font(hex_color="FFFFFF", bold=True, size=10):
    return Font(name="Calibri", color=hex_color, bold=bold, size=size)

def thin_border():
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def thick_border():
    return Border(left=thick, right=thick, top=thick, bottom=thick)

def center(wrap=True):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def left(wrap=True):
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)


# ─────────────────────────────────────────────
# SHEET 1 — Main Reference
# ─────────────────────────────────────────────
ws = wb.active
ws.title = "Backend → Engine Data Map"

# Column widths
col_widths = [5, 26, 22, 14, 40, 30, 22]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.row_dimensions[1].height = 36
ws.row_dimensions[2].height = 20

# ── TITLE ROW ──────────────────────────────
ws.merge_cells("A1:G1")
c = ws["A1"]
c.value        = "Simple Saturation Analysis  —  Backend → Calculation Engine Data Reference"
c.fill         = hdr_fill(CLR_HEADER_DARK)
c.font         = Font(name="Calibri", color="FFFFFF", bold=True, size=14)
c.alignment    = center()

# ── SUB-HEADER ROW ─────────────────────────
headers = ["#", "Parameter / Field", "JSON   Key", "Data Type", "Description / Note", "Where Used in Engine", "Source"]
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=2, column=col)
    cell.value     = h
    cell.fill      = hdr_fill(CLR_HEADER_MID)
    cell.font      = hdr_font()
    cell.alignment = center()
    cell.border    = thin_border()

# ─────────────────────────────────────────────
# DATA ROWS
# ─────────────────────────────────────────────
# Format: (section_label, section_color, label_color, rows)
# Each row: (no, param, json_key, dtype, description, engine_use, source)

SECTION_A_ROWS = [
    ("1",  "Calcium",            "Ca",          "float (mg/L)",  "Ca²⁺ concentration from lab report",                        "base_water_params['Ca'] × CoC → PHREEQC input",                 "Backend DB\n(water_analyses table)"),
    ("2",  "Magnesium",          "Mg",          "float (mg/L)",  "Mg²⁺ concentration",                                         "base_water_params['Mg'] × CoC → PHREEQC input",                 "Backend DB"),
    ("3",  "Sodium",             "Na",          "float (mg/L)",  "Na⁺ — also adjusted if NaOH used for pH",                   "base_water_params['Na'] × CoC ± pH_adjustment → PHREEQC",       "Backend DB"),
    ("4",  "Potassium",          "K",           "float (mg/L)",  "K⁺ concentration",                                          "base_water_params['K'] × CoC → PHREEQC",                        "Backend DB"),
    ("5",  "Bicarbonate (Alk)",  "HCO3",        "float (mg/L)",  "HCO₃⁻ / Alkalinity as CaCO₃; adjusted by pH chemical",     "base_water_params['HCO3'] × CoC ± pH_adjustment → PHREEQC",     "Backend DB"),
    ("6",  "Sulfate",            "SO4",         "float (mg/L)",  "SO₄²⁻; increases if H₂SO₄ used for pH",                    "base_water_params['SO4'] × CoC ± pH_adjustment → PHREEQC",      "Backend DB"),
    ("7",  "Chloride",           "Cl",          "float (mg/L)",  "Cl⁻; increases if HCl used for pH",                        "base_water_params['Cl'] × CoC ± pH_adjustment → PHREEQC",       "Backend DB"),
    ("8",  "Silica",             "SiO2",        "float (mg/L)",  "Dissolved SiO₂",                                            "base_water_params['SiO2'] × CoC → PHREEQC",                     "Backend DB"),
    ("9",  "Barium",             "Ba",          "float (mg/L)",  "Ba²⁺ (for Barite SI)",                                      "base_water_params['Ba'] × CoC → PHREEQC",                       "Backend DB"),
    ("10", "Strontium",          "Sr",          "float (mg/L)",  "Sr²⁺ (for Celestite SI)",                                   "base_water_params['Sr'] × CoC → PHREEQC",                       "Backend DB"),
    ("11", "Iron",               "Fe",          "float (mg/L)",  "Fe²⁺ total iron",                                           "base_water_params['Fe'] × CoC → PHREEQC",                       "Backend DB"),
    ("12", "pH (base value)",    "pH",          "float",         "Original lab pH.\nIF user selects 'Fixed pH' → overridden", "If natural → used as-is\nIf fixed → replaced by user value",   "Backend DB"),
    ("13", "Temperature (base)", "Temperature", "float (°C)",    "Original water temp in °C",                                 "Base reference — user may override via temp range",              "Backend DB"),
]

SECTION_B_ROWS = [
    ("1",  "salt_id",            "salt_id",           "string",           "Which salt to analyze (e.g. 'Calcite')",                    "Filters SI output from PHREEQC result list",                    "Backend DB\n(salt list table)"),
    ("2",  "Max SI at Dose",     "max_si_at_dose",    "float",            "Threshold SI for selected product at selected dose (ppm)",  "Color coding: SI < this → 🟢 GREEN",                           "Backend DB\n(rm_salt_thresholds)"),
    ("3",  "Yellow Cushion Min", "yellow_cushion_min","float",            "Lower bound of yellow zone",                                "Color coding: SI ≥ this → starts 🟡 YELLOW zone",             "Backend DB\n(rm_salt_thresholds)"),
    ("4",  "Yellow Cushion Max", "yellow_cushion_max","float",            "Upper bound of yellow zone / max protection limit",        "Color coding: SI > this → 🔴 RED",                             "Backend DB\n(rm_salt_thresholds)"),
    ("5",  "Treatment Name",     "treatment_name",    "string",           "Product name (e.g. HEDP)",                                  "Display label only",                                            "Backend DB\n(products table)"),
    ("6",  "Raw Material Name",  "raw_material_name", "string",           "Active raw material in product",                            "Used to fetch correct threshold row",                           "Backend DB\n(raw_materials table)"),
]

SECTION_C_ROWS = [
    ("1",  "CoC Min",            "coc_min",           "float",            "Minimum Cycles of Concentration",                          "grid_points loop start",                                        "User Input (Frontend)"),
    ("2",  "CoC Max",            "coc_max",           "float",            "Maximum Cycles of Concentration",                          "grid_points loop end",                                          "User Input"),
    ("3",  "CoC Interval",       "coc_interval",      "float",            "Step size between CoC values",                             "Loop step: range(coc_min, coc_max, interval)",                  "User Input"),
    ("4",  "Temp Min (F or C)", "temp_min",           "float",            "Minimum temperature — convert to °C before PHREEQC",       "grid_points[i]['temp']",                                        "User Input"),
    ("5",  "Temp Max (F or C)", "temp_max",           "float",            "Maximum temperature",                                      "grid_points[i]['temp']",                                        "User Input"),
    ("6",  "Temp Interval",      "temp_interval",     "float",            "Step size between temperature values",                     "Loop step",                                                     "User Input"),
    ("7",  "pH Mode",            "ph_mode",           "string",           "'fixed' or 'natural'",                                     "If fixed: override pH + apply chemical adjustment",             "User Input"),
    ("8",  "Fixed pH Value",     "fixed_ph",          "float",            "pH value if mode = 'fixed' (e.g. 8.2)",                    "grid_points[i]['pH']",                                          "User Input"),
    ("9",  "pH Chemical",        "adjustment_chemical","string",           "'H2SO4' / 'HCl' / 'NaOH'",                                "Determines which ions to adjust (Alk, SO4, Cl, or Na)",        "User Input"),
    ("10", "Dosage (ppm)",       "dosage_ppm",         "float",           "Treatment product dose in ppm",                            "Used to look up max_si_at_dose from thresholds",               "User Input"),
    ("11", "Balance Cation",     "balance_cation",    "string",           "'Na' or 'K' — for ion balance correction",                 "Passed to ion_balance() method",                               "User Input"),
    ("12", "Balance Anion",      "balance_anion",     "string",           "'Cl' or 'SO4' — for ion balance correction",               "Passed to ion_balance() method",                               "User Input"),
]

SECTION_D_ROWS = [
    ("1",  "Saturation Index",   "saturation_indices","dict",             "SI for every salt PHREEQC calculates",                     "Filter by selected salt_id for graph Y-axis",                   "PHREEQC Output"),
    ("2",  "Ionic Strength",     "ionic_strength",    "float",            "IS of solution at this grid point",                        "Used to auto-select phreeqc.dat vs pitzer.dat",                 "PHREEQC Output"),
    ("3",  "Charge Balance %",   "charge_balance_error_pct","float",      "% error in charge balance",                                "Logged; ion_balance() corrects if > ±5%",                       "PHREEQC Output"),
    ("4",  "Color Code",         "color_code",        "string",           "'green' / 'yellow' / 'red'",                               "3D graph bar coloring",                                         "Calculation Engine"),
    ("5",  "Grid Point: CoC",    "_grid_CoC",         "float",            "CoC value for this result row",                            "3D Graph X or Z axis",                                          "Calculation Engine"),
    ("6",  "Grid Point: Temp",   "_grid_temp",        "float",            "Temperature (°C) for this result row",                     "3D Graph X or Z axis",                                          "Calculation Engine"),
    ("7",  "Grid Point: pH",     "_grid_pH",          "float",            "pH for this result row",                                   "3D Graph X or Z axis (if pH is ranged)",                        "Calculation Engine"),
]

def write_section_header(ws, row, label, color, font_color):
    ws.merge_cells(f"A{row}:G{row}")
    c = ws.cell(row=row, column=1)
    c.value     = label
    c.fill      = row_fill(color)
    c.font      = Font(name="Calibri", color=font_color, bold=True, size=10)
    c.alignment = left()
    ws.row_dimensions[row].height = 18

def write_data_rows(ws, start_row, rows, fill_color, alt=True):
    for i, (no, param, key, dtype, desc, engine, source) in enumerate(rows):
        r = start_row + i
        ws.row_dimensions[r].height = 40
        fill = row_fill(fill_color) if (not alt or i % 2 == 0) else row_fill(CLR_ROW_ALT)
        data = [no, param, key, dtype, desc, engine, source]
        for col, val in enumerate(data, 1):
            cell = ws.cell(row=r, column=col)
            cell.value     = val
            cell.fill      = fill
            cell.font      = Font(name="Calibri", size=9)
            cell.alignment = left() if col > 2 else center(wrap=True)
            cell.border    = thin_border()

# ── SECTION A: Backend DB Water Parameters ──
write_section_header(ws, 3,
    "  SECTION A — Water Parameters  (Backend DB → Calculation Engine)",
    "1F3864", "FFFFFF")
write_data_rows(ws, 4, SECTION_A_ROWS, CLR_SECTION_A)

# ── SECTION B: Backend DB Thresholds ────────
r_b = 4 + len(SECTION_A_ROWS)
write_section_header(ws, r_b,
    "  SECTION B — Treatment / Salt Threshold Data  (Backend DB → Color Coding Logic)",
    "375623", "FFFFFF")
write_data_rows(ws, r_b + 1, SECTION_B_ROWS, CLR_SECTION_B)

# ── SECTION C: User Input ───────────────────
r_c = r_b + 1 + len(SECTION_B_ROWS)
write_section_header(ws, r_c,
    "  SECTION C — User Input  (Frontend → Calculation Engine — NOT from Backend)",
    "7F6000", "FFFFFF")
write_data_rows(ws, r_c + 1, SECTION_C_ROWS, CLR_SECTION_C)

# ── SECTION D: PHREEQC Output ───────────────
r_d = r_c + 1 + len(SECTION_C_ROWS)
write_section_header(ws, r_d,
    "  SECTION D — PHREEQC Output + Engine Results  (Calculated — stored in DB)",
    "833C00", "FFFFFF")
write_data_rows(ws, r_d + 1, SECTION_D_ROWS, CLR_SECTION_D)


# ─────────────────────────────────────────────
# SHEET 2 — pH Adjustment Cheat Sheet
# ─────────────────────────────────────────────
ws2 = wb.create_sheet("pH Adjustment Rules")
for i, w in enumerate([5, 22, 30, 34, 34], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

ws2.merge_cells("A1:E1")
c = ws2["A1"]
c.value     = "pH Adjustment Chemical — Ion Correction Rules"
c.fill      = hdr_fill(CLR_HEADER_DARK)
c.font      = Font(name="Calibri", color="FFFFFF", bold=True, size=13)
c.alignment = center()
ws2.row_dimensions[1].height = 32

hdrs2 = ["#", "Chemical", "Full Name", "Alkalinity (HCO3) Effect", "Counter-Ion Effect"]
for col, h in enumerate(hdrs2, 1):
    cell = ws2.cell(row=2, column=col)
    cell.value     = h
    cell.fill      = hdr_fill(CLR_HEADER_MID)
    cell.font      = hdr_font()
    cell.alignment = center()
    cell.border    = thin_border()
ws2.row_dimensions[2].height = 20

ph_rows = [
    ("1", "H₂SO₄", "Sulfuric Acid",
     "REDUCES Alkalinity\n1 ppm H₂SO₄  →  HCO3 decreases by 1.00 ppm",
     "INCREASES Sulfate\n1 ppm H₂SO₄  →  SO4 increases by 1.00 ppm"),
    ("2", "HCl",   "Hydrochloric Acid",
     "REDUCES Alkalinity\n1 ppm HCl  →  HCO3 decreases by 1.37 ppm",
     "INCREASES Chloride\n1 ppm HCl  →  Cl increases by 0.97 ppm"),
    ("3", "NaOH",  "Sodium Hydroxide",
     "INCREASES Alkalinity\n1 ppm NaOH  →  HCO3 increases by 1.25 ppm",
     "INCREASES Sodium\n1 ppm NaOH  →  Na increases by 0.57 ppm"),
]

fills2 = ["D6E4F7", "E2EFDA", "FCE4D6"]
for i, (no, chem, name, alk, counter) in enumerate(ph_rows):
    r = 3 + i
    ws2.row_dimensions[r].height = 50
    fl = row_fill(fills2[i])
    for col, val in enumerate([no, chem, name, alk, counter], 1):
        cell = ws2.cell(row=r, column=col)
        cell.value     = val
        cell.fill      = fl
        cell.font      = Font(name="Calibri", size=10, bold=(col == 2))
        cell.alignment = center() if col <= 3 else left()
        cell.border    = thin_border()

# Formula display
ws2.row_dimensions[7].height = 18
ws2.merge_cells("A7:E7")
ws2["A7"].value     = "Formula applied BEFORE CoC multiplication:   new_ion_value = base_ion_value ± (dose_of_chemical × factor)"
ws2["A7"].font      = Font(name="Calibri", size=9, italic=True, color="595959")
ws2["A7"].alignment = left()


# ─────────────────────────────────────────────
# SHEET 3 — Flow Diagram (text)
# ─────────────────────────────────────────────
ws3 = wb.create_sheet("Data Flow Summary")
ws3.column_dimensions["A"].width = 90
ws3.row_dimensions[1].height = 32

ws3.merge_cells("A1:A1")
ws3["A1"].value     = "Data Flow:  Backend  →  Calculation Engine  →  Output"
ws3["A1"].fill      = hdr_fill(CLR_HEADER_DARK)
ws3["A1"].font      = Font(name="Calibri", color="FFFFFF", bold=True, size=13)
ws3["A1"].alignment = center()

flow_lines = [
    "",
    "STEP 1 — Frontend calls: GET /api/saturation/water-analysis/{analysis_id}",
    "         Backend returns: Ca, Mg, Na, K, HCO3, SO4, Cl, SiO2, Ba, Sr, Fe, pH, Temperature  (all in mg/L or °C)",
    "",
    "STEP 2 — Frontend calls: GET /api/saturation/treatments/{product_id}?salt={salt_id}&dose={dosage_ppm}",
    "         Backend returns: max_si_at_dose, yellow_cushion_min, yellow_cushion_max",
    "",
    "STEP 3 — User inputs (from UI, NOT backend):",
    "         • CoC range: min=1, max=10, interval=1",
    "         • Temperature range: min=110°F, max=160°F, interval=10°F",
    "         • pH mode: 'fixed' at 8.2  with  H2SO4",
    "         • Salt: Calcite",
    "         • Treatment: HEDP  at  2 ppm",
    "",
    "STEP 4 — Calculation Engine applies pH adjustment to ions FIRST:",
    "         (for each CoC point, before multiplying):",
    "         HCO3_adjusted = HCO3_base  -  (H2SO4_dose × 1.0)",
    "         SO4_adjusted  = SO4_base   +  (H2SO4_dose × 1.0)",
    "",
    "STEP 5 — Calculation Engine builds grid_points list:",
    "         For CoC in [1,2,3,...,10]:",
    "           For Temp in [43.3, 48.9, 54.4, 60.0, 65.6, 71.1]°C:   ← converted from °F",
    "             grid_points.append({ 'pH': 8.2, 'CoC': CoC, 'temp': Temp })",
    "         → Total: 10 × 6 = 60 grid points",
    "",
    "STEP 6 — Calculation Engine calls existing PHREEQCService:",
    "         phreeqc_service.select_database(...)   → picks phreeqc.dat or pitzer.dat",
    "         phreeqc_service.ion_balance(...)        → corrects charge balance",
    "         phreeqc_service.run_batch_solution_spread(base_params, grid_points, db)",
    "         → Returns: SI values for all salts at all 60 grid points",
    "",
    "STEP 7 — Color coding (using Section B thresholds):",
    "         For each result point, get SI for 'Calcite':",
    "           SI < max_si_at_dose_2ppm  → GREEN",
    "           yellow_min ≤ SI ≤ yellow_max  → YELLOW",
    "           SI > yellow_max  → RED",
    "",
    "STEP 8 — Store results in DB (saturation_run_results table)",
    "         Send to frontend → display 3D graph",
]

for i, line in enumerate(flow_lines, 2):
    ws3.row_dimensions[i].height = 16
    c = ws3.cell(row=i, column=1)
    c.value = line
    is_step = line.startswith("STEP")
    c.font  = Font(name="Consolas", size=9,
                   bold=is_step,
                   color=CLR_HEADER_DARK if is_step else "1A1A1A")
    c.alignment = left(wrap=False)


# ─────────────────────────────────────────────
# SAVE
# ─────────────────────────────────────────────
out_path = r"c:\Users\shaikat\Desktop\jimgreen-Ai-Backend - new feature added-final\backend\Saturation_Analysis_Data_Reference.xlsx"
wb.save(out_path)
print(f"✅ Excel saved: {out_path}")
